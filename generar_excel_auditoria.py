# -*- coding: utf-8 -*-
"""
Generador de Checklist de Auditoría de Salones CCTV en Excel
Con formato, colores y formato condicional
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def crear_auditoria_excel():
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría Salones"
    
    # Definir encabezados
    headers = [
        "Nº", "Abonado", "Cliente", "Ubicación", "Estado Visionado",
        "IP/Dominio", "Puerto HTTP", "Puerto RTSP", "Puerto Servidor", "P2P ID",
        "Usuario", "Contraseña", "Modelo DVR", "Nº Cámaras Total", "Nº Cámaras OK",
        "Software CRA", "Tel. Responsable", "Tel. Informático", "Proveedor Internet",
        "Velocidad (Mbps)", "Última Conexión", "Problemas Detectados", 
        "Acciones Requeridas", "Prioridad", "Garantía", "Fin Garantía", "Completado"
    ]
    
    # Escribir encabezados
    ws.append(headers)
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Añadir datos de los 58 salones
    for i in range(1, 59):
        row_data = [
            i,  # Nº
            "",  # Abonado
            "",  # Cliente
            "",  # Ubicación
            "No visible" if i != 17 and i != 43 else "Visible",  # Estado
            "",  # IP/Dominio
            "",  # Puerto HTTP
            "",  # Puerto RTSP
            "",  # Puerto Servidor
            "",  # P2P ID
            "admin",  # Usuario
            "",  # Contraseña
            "",  # Modelo DVR
            "",  # Nº Cámaras Total
            "",  # Nº Cámaras OK
            "",  # Software CRA
            "",  # Tel. Responsable
            "",  # Tel. Informático
            "",  # Proveedor Internet
            "",  # Velocidad
            "",  # Última Conexión
            "",  # Problemas
            "",  # Acciones
            "Media" if i != 17 and i != 43 else "Baja",  # Prioridad
            "No" if i != 17 and i != 43 else "Sí",  # Garantía
            "",  # Fin Garantía
            "No" if i != 17 and i != 43 else "Sí"  # Completado
        ]
        ws.append(row_data)
    
    # Definir colores de fondo para grupos de columnas
    # Grupo 1: Info Básica (A-E) - Verde claro
    fill_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    # Grupo 2: Red (F-J) - Azul claro
    fill_blue = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    # Grupo 3: Acceso (K-P) - Amarillo claro
    fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    # Grupo 4: Contactos (Q-S) - Naranja claro
    fill_orange = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    # Grupo 5: Seguimiento (T-AA) - Morado claro
    fill_purple = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    
    # Aplicar colores a las filas de datos (2-59)
    for row in range(2, 60):
        # Grupo 1: A-E (columnas 1-5)
        for col in range(1, 6):
            if row % 2 == 0:  # Filas pares
                ws.cell(row=row, column=col).fill = fill_green
        
        # Grupo 2: F-J (columnas 6-10)
        for col in range(6, 11):
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_blue
        
        # Grupo 3: K-P (columnas 11-16)
        for col in range(11, 17):
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_yellow
        
        # Grupo 4: Q-S (columnas 17-19)
        for col in range(17, 20):
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_orange
        
        # Grupo 5: T-AA (columnas 20-27)
        for col in range(20, 28):
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = fill_purple
    
    # Formato condicional para "Estado Visionado" (columna E)
    # Verde para "Visible"
    fill_visible = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    font_visible = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('E2:E59',
        CellIsRule(operator='containsText', formula=['"Visible"'], 
                   fill=fill_visible, font=font_visible))
    
    # Rojo para "No visible"
    fill_no_visible = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    font_no_visible = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('E2:E59',
        CellIsRule(operator='containsText', formula=['"No visible"'], 
                   fill=fill_no_visible, font=font_no_visible))
    
    # Naranja para "Intermitente"
    fill_intermitente = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    font_intermitente = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('E2:E59',
        CellIsRule(operator='containsText', formula=['"Intermitente"'], 
                   fill=fill_intermitente, font=font_intermitente))
    
    # Formato condicional para "Prioridad" (columna X)
    # Crítica - Rojo oscuro
    fill_critica = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    font_critica = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('X2:X59',
        CellIsRule(operator='equal', formula=['"Crítica"'], 
                   fill=fill_critica, font=font_critica))
    
    # Alta - Naranja
    fill_alta = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    font_alta = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('X2:X59',
        CellIsRule(operator='equal', formula=['"Alta"'], 
                   fill=fill_alta, font=font_alta))
    
    # Media - Amarillo
    fill_media = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")
    font_media = Font(color="000000")
    ws.conditional_formatting.add('X2:X59',
        CellIsRule(operator='equal', formula=['"Media"'], 
                   fill=fill_media, font=font_media))
    
    # Baja - Verde
    fill_baja = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
    font_baja = Font(color="FFFFFF")
    ws.conditional_formatting.add('X2:X59',
        CellIsRule(operator='equal', formula=['"Baja"'], 
                   fill=fill_baja, font=font_baja))
    
    # Formato condicional para "Completado" (columna AA)
    # Sí - Verde
    fill_completado = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    font_completado = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add('AA2:AA59',
        CellIsRule(operator='equal', formula=['"Sí"'], 
                   fill=fill_completado, font=font_completado))
    
    # No - Gris
    fill_no_completado = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    font_no_completado = Font(color="666666")
    ws.conditional_formatting.add('AA2:AA59',
        CellIsRule(operator='equal', formula=['"No"'], 
                   fill=fill_no_completado, font=font_no_completado))
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 5,   # Nº
        'B': 12,  # Abonado
        'C': 20,  # Cliente
        'D': 25,  # Ubicación
        'E': 15,  # Estado
        'F': 20,  # IP/Dominio
        'G': 12,  # Puerto HTTP
        'H': 12,  # Puerto RTSP
        'I': 14,  # Puerto Servidor
        'J': 20,  # P2P ID
        'K': 10,  # Usuario
        'L': 12,  # Contraseña
        'M': 20,  # Modelo DVR
        'N': 14,  # Nº Cámaras Total
        'O': 14,  # Nº Cámaras OK
        'P': 15,  # Software CRA
        'Q': 15,  # Tel. Responsable
        'R': 15,  # Tel. Informático
        'S': 16,  # Proveedor
        'T': 12,  # Velocidad
        'U': 15,  # Última Conexión
        'V': 30,  # Problemas
        'W': 30,  # Acciones
        'X': 10,  # Prioridad
        'Y': 10,  # Garantía
        'Z': 12,  # Fin Garantía
        'AA': 12  # Completado
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Altura de fila de encabezado
    ws.row_dimensions[1].height = 30
    
    # Centrar texto en columnas específicas
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, 60):
        for col in [1, 5, 7, 8, 9, 11, 14, 15, 16, 20, 24, 25, 27]:  # Columnas a centrar
            ws.cell(row=row, column=col).alignment = center_alignment
    
    # Ajuste de texto para columnas de problemas y acciones
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(2, 60):
        ws.cell(row=row, column=22).alignment = wrap_alignment  # Problemas
        ws.cell(row=row, column=23).alignment = wrap_alignment  # Acciones
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    # Aplicar filtros
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}59"
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for row in range(1, 60):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Guardar archivo
    output_file = r"c:\Users\Juan Luis\.gemini\antigravity\scratch\Alarmas Gama\Auditoria_Salones_CCTV.xlsx"
    wb.save(output_file)
    print(f"✅ Archivo Excel creado exitosamente: {output_file}")
    print(f"📊 Total de salones: 58")
    print(f"🎨 Formato aplicado con 5 grupos de colores")
    print(f"✨ Formato condicional activado para Estado, Prioridad y Completado")

if __name__ == "__main__":
    crear_auditoria_excel()
